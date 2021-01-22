import os

import openpyxl
class reuseable:

    # def get_driver_path(self):
    #     cwd = os.path.dirname(__file__)
    #     driver_path = os.path.join(cwd, 'Driver/chromedriver')
    #     return driver_path

    def get_download_dir(self):
        cwd = os.path.dirname(__file__)
        download_path = os.path.join(cwd, '/home/chetan/Downloads/Atria_Project/AtriaTesting-1.0/Downloads')
        return download_path

    def getRowCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return(sheet.max_row)

    def getColumnCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    def readData(file,sheetName,rownum,columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file,sheetName,rownum,columnno,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)
